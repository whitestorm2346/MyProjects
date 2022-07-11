import os
import urllib.request as req
import bs4
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class TravelAgencyInfo:
    def __init__(self, title, title_eng, class_, phone, address, member, insurance_info):
        self.title = title
        self.title_eng = title_eng
        self.class_ = class_
        self.phone = phone
        self.address = address
        self.member = member
        self.insurance_info = insurance_info

travel_agency_class = {
    '甲種': '%e7%94%b2%e7%a8%ae',
    '乙種': '%e4%b9%99%e7%a8%ae',
    '綜合': '%e7%b6%9c%e5%90%88'
}

try:
    with open('旅行社資料.xlsx', 'r'):
        pass
except FileNotFoundError:
    wb = Workbook()
    file_path = os.path.dirname(os.path.abspath(__file__))
    base_path = os.path.join(file_path, '旅行社資料.xlsx')

    wb.save('旅行社資料.xlsx')

wb = load_workbook('旅行社資料.xlsx')

for sheet in wb.sheetnames:
    wb.remove(wb[sheet])

for ta_class in travel_agency_class:
    page_count = 1
    print(f'開始抓取{ta_class}旅行社的資料')

    wb.create_sheet(title=ta_class + '旅行社')
    ws = wb.active

    ws.append([
        '旅行社',
        'Travel Agency',
        '公司類別',
        '電話',
        '地址',
        '品保會員',
        '保險資料'
    ])

    for col in range(1, 7):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)

    while True:
        url = 'https://admin.taiwan.net.tw/TravelIndustryList.aspx?' \
              f'Pindex={page_count}&SIKEY={travel_agency_class[ta_class]}'
        request = req.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/102.0.5005.124 Safari/537.36 Edg/102.0.1245.41"
        })

        with req.urlopen(request) as response:
            code = response.read().decode('utf-8')
            root = bs4.BeautifulSoup(code, "html.parser")

            travel_agency = root.find_all("div", class_="travel_item")

            if not travel_agency:
                break

            for ta in travel_agency:
                travel_agency_a = ta.find('a', href=True)
                request2 = req.Request('https://admin.taiwan.net.tw/' + travel_agency_a['href'])

                with req.urlopen(request2) as response2:
                    code2 = response2.read().decode('utf-8')
                    root2 = bs4.BeautifulSoup(code2, "html.parser")

                    promote_tit = root2.find("h4", class_="promote_tit")
                    title = promote_tit.string.split(maxsplit=1)

                    if len(title) == 1:
                        title.append('NONE')

                travel_data = ta.find("div", class_="travel_data")
                travel_data_list = travel_data.find_all("dd")

                ws.append([
                    title[0],
                    title[1],
                    travel_data_list[0].string,
                    travel_data_list[1].string,
                    travel_data_list[2].string,
                    travel_data_list[3].string,
                    travel_data_list[4].string
                ])

                print('旅行社: {0}'.format(title[0]))
                print('Travel Agency: {0}'.format(title[1]))
                print('公司類別: {0}'.format(travel_data_list[0].string))
                print('電話: {0}'.format(travel_data_list[1].string))
                print('地址: {0}'.format(travel_data_list[2].string))
                print('品保會員: {0}'.format(travel_data_list[3].string))
                print('保險資料: {0}'.format(travel_data_list[4].string), end='\n\n')

        page_count += 1
    
    print('完成儲存' + ta_class + '旅行社的資料')

wb.save('旅行社資料.xlsx')
print('已將所有資料儲存於 旅行社資料.xlsx')

os.system("pause")


